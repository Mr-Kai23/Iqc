import openpyxl
import xlrd
from openpyxl.styles import Side, Border, fills
import pandas as pd


class make_sheet(object):

	def __init__(self, xfile, fname, result_list):

		self.xfile = xfile
		self.fname = fname
		self.result_list = result_list

	def top_module_sum(self):
		wb = openpyxl.load_workbook(self.xfile)

		ws = wb.active
		# --------------header---------
		for i in range(1, len(self.result_list) + 1):
			for j in range(min(len(self.result_list[0]), len(self.result_list[2]))):
				cell = ws.cell(row=i + 4, column=j + 2)

				# 添加汇总信息颜色标记
				if self.result_list[i - 1][-1] == True:
					cell.fill = fills.GradientFill(stop=('FFE4B5', 'FFE4B5'))
				cell.value = self.result_list[i - 1][j]

		wb.save(self.fname)

	def Other_module_sum(self):
		wb = openpyxl.load_workbook(self.xfile)

		ws = wb.active

		for i in range(1, len(self.result_list) + 1):
			for j in range(min(len(self.result_list[0]), len(self.result_list[2]))):
				cell = ws.cell(row=i + 3, column=j + 2)

				# 添加汇总信息颜色标记
				if self.result_list[i - 1][-1] == True:
					cell.fill = fills.GradientFill(stop=('FFE4B5', 'FFE4B5'))

				cell.value = self.result_list[i - 1][j]

		wb.save(self.fname)

	def ttl_sum_top(self):

		data = xlrd.open_workbook(self.fname)
		wb = openpyxl.load_workbook(self.fname)
		worksheet = wb.active

		left, right, top, bottom = [Side(style='thin', color='000000')]*4
		border = Border(left=left, right=right, top=top, bottom=bottom)

		for row in worksheet.iter_rows(min_row=2, min_col=2, max_row=4):
			for cell in row:
				cell.border = border		

		sheet_list = data.sheet_names()
		table = data.sheet_by_name(sheet_list[0])
		nrows = table.nrows

		sum_ttl = ['O', 'P', 'Q', 'R', 'T', 'U', 'V', 'Y', 'Z', 'AA', 'AD', 'AE', 'AF', 'AH', 'AJ', 'AL', 'AN', 'AP',
		'AQ', 'AR', 'AT', 'AU', 'AV', 'AX', 'AY', 'AZ', 'BB', 'BC', 'BD']

		# sep_ttl = ['S', 'W', 'AB', 'AS', 'AW', 'BA', 'BE', 'AG', 'AI', 'AK', 'AM', 'AO']

		worksheet.cell(row=nrows + 1, column=13, value='TTL:')
		newList = []

		for result in self.result_list:
			smaList = []
			smaList.append(result[4])
			smaList.append(result[12])
			smaList.append(result[13])
			smaList.append(result[13])
			newList.append(smaList)

		newData = pd.DataFrame(newList).groupby([0, 1, 2])
		newValue = sum(newData.mean()[3].values.tolist())

		for ttl_su in sum_ttl:
			if ttl_su == 'O':
				worksheet['O' + str(nrows + 1)] = newValue
			else:
				worksheet[ttl_su + str(nrows + 1)] = '=SUM(' + ttl_su + '4:' + ttl_su + str(nrows) + ')'

		wb.save(self.fname)

	def ttl_sum_other(self, flag):

		data = xlrd.open_workbook(self.fname)
		wb = openpyxl.load_workbook(self.fname)
		worksheet = wb.active

		left, right, top, bottom = [Side(style='thin', color='000000')]*4

		border = Border(left=left, right=right, top=top, bottom=bottom)

		for row in worksheet.iter_rows(min_row=2, min_col=2, max_row=4):
			for cell in row:
				cell.border = border	

		sheet_list = data.sheet_names()
		table = data.sheet_by_name(sheet_list[0])
		nrows = table.nrows

		if flag:

			sum_ttl = ['N', 'O', 'P', 'Q', 'S', 'T', 'U', 'X', 'Y', 'Z', 'AC', 'AD', 'AE', 'AH', 'AI', 'AJ']
		else:
			sum_ttl = ['N', 'O', 'P', 'Q', 'S', 'T', 'U', 'X', 'Y', 'Z', 'AC', 'AD', 'AE']

		worksheet.cell(row=nrows+1, column=12, value='TTL:')

		# pn、config、eee相同的数据不相加
		newList = []

		for result in self.result_list:
			smaList = []
			smaList.append(result[5])
			smaList.append(result[11])
			smaList.append(result[12])
			smaList.append(result[12])
			newList.append(smaList)
			newData = pd.DataFrame(newList).groupby([0, 1, 2])
			newValue = sum(newData.mean()[3].values.tolist())

		for ttl_su in sum_ttl:
			if ttl_su == 'N':
				worksheet['N' + str(nrows + 1)] = newValue
			else:
				worksheet[ttl_su + str(nrows + 1)] = '=SUM(' + ttl_su + '4:' + ttl_su + str(nrows) + ')'

		wb.save(self.fname)





